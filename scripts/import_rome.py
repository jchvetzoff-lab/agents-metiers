#!/usr/bin/env python3
"""
Script d'import du référentiel ROME depuis les fichiers XLSX téléchargés sur data.gouv.fr.
Importe les 1584 fiches métiers, leurs appellations, compétences et savoirs.
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from datetime import datetime
from database.repository import Repository
from database.models import (
    FicheMetier, SalairesMetier, PerspectivesMetier,
    MetadataFiche, TendanceMetier, StatutFiche, Base
)
from config import get_config


# ===========================================================================
# Parsing des fichiers XLSX
# ===========================================================================

def parse_nom_genre(nom_complet: str) -> dict:
    """
    Parse un nom au format 'Masculin / Féminin complément' du ROME.

    Exemples:
        'Développeur / Développeuse informatique' -> masc='Développeur informatique', fem='Développeuse informatique'
        'Tractoriste agricole' -> masc=fem=epicene='Tractoriste agricole'
        'Aide-soignant / Aide-soignante' -> masc='Aide-soignant', fem='Aide-soignante'
    """
    if ' / ' not in nom_complet:
        return {
            'masculin': nom_complet,
            'feminin': nom_complet,
            'epicene': nom_complet
        }

    parts = nom_complet.split(' / ', 1)
    masculin_part = parts[0].strip()
    feminin_rest = parts[1].strip()

    # Le complément est la partie après le mot féminin
    # Ex: "Développeur / Développeuse informatique"
    # masculin_part = "Développeur"
    # feminin_rest = "Développeuse informatique"
    # On doit extraire: fem_mot = "Développeuse", complement = "informatique"

    # Compter les mots du masculin pour savoir combien de mots forme le mot féminin
    masc_words = masculin_part.split()
    fem_words = feminin_rest.split()

    # Le mot féminin a le même nombre de mots que le masculin
    nb_mots_nom = len(masc_words)

    if len(fem_words) > nb_mots_nom:
        fem_nom = ' '.join(fem_words[:nb_mots_nom])
        complement = ' '.join(fem_words[nb_mots_nom:])
        masculin = f"{masculin_part} {complement}"
        feminin = f"{fem_nom} {complement}"
    else:
        masculin = masculin_part
        feminin = feminin_rest

    # Forme épicène : utilise la forme la plus inclusive
    epicene = nom_complet

    return {
        'masculin': masculin,
        'feminin': feminin,
        'epicene': epicene
    }


def load_arborescence_principale(filepath: Path) -> dict:
    """
    Charge l'arborescence principale du ROME.

    Retourne un dict avec:
    - domaines: {lettre: nom}
    - sous_domaines: {code: nom}
    - fiches: {code_rome: {nom, domaine, sous_domaine, appellations, code_ogr}}
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[1]]  # Deuxième feuille = données

    domaines = {}
    sous_domaines = {}
    fiches = {}

    current_domaine = ""
    current_domaine_nom = ""
    current_sous_domaine = ""
    current_sous_domaine_nom = ""

    for row in ws.iter_rows(min_row=2, values_only=True):
        col1 = str(row[0]).strip() if row[0] else ""
        col2 = str(row[1]).strip() if row[1] else ""
        col3 = str(row[2]).strip() if row[2] else ""
        col4 = str(row[3]).strip() if row[3] else ""
        col5 = str(row[4]).strip() if row[4] else ""

        # Grand domaine (A, B, C...)
        if col1 and not col2 and not col3 and col4:
            current_domaine = col1
            current_domaine_nom = col4
            domaines[col1] = col4
            continue

        # Sous-domaine (A11, A12...)
        if col1 and col2 and not col3 and col4:
            current_sous_domaine = f"{col1}{col2}"
            current_sous_domaine_nom = col4
            sous_domaines[current_sous_domaine] = col4
            continue

        # Fiche ou appellation
        if col3:
            code_rome = f"{col1}{col2}{col3}"

            if code_rome not in fiches:
                # Première entrée = nom principal de la fiche
                noms = parse_nom_genre(col4)
                fiches[code_rome] = {
                    'nom_complet': col4,
                    'nom_masculin': noms['masculin'],
                    'nom_feminin': noms['feminin'],
                    'nom_epicene': noms['epicene'],
                    'code_ogr': col5,
                    'domaine': current_domaine_nom,
                    'domaine_code': current_domaine,
                    'sous_domaine': current_sous_domaine_nom,
                    'sous_domaine_code': current_sous_domaine,
                    'appellations': []
                }
            else:
                # Entrées suivantes = appellations alternatives
                fiches[code_rome]['appellations'].append({
                    'nom': col4,
                    'code_ogr': col5
                })

    wb.close()
    return {
        'domaines': domaines,
        'sous_domaines': sous_domaines,
        'fiches': fiches
    }


def load_competences(filepath: Path) -> list:
    """Charge les macro-compétences depuis le fichier XLSX."""
    wb = openpyxl.load_workbook(filepath)

    # Chercher la feuille de données brutes macro-compétences
    sheet_name = None
    for name in wb.sheetnames:
        if 'brutes' in name.lower() and 'macro' in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    competences = []
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4]:  # Macro-compétence
            comp = str(row[4]).strip()
            if comp not in seen:
                seen.add(comp)
                competences.append({
                    'domaine': str(row[1]).strip() if row[1] else '',
                    'enjeu': str(row[2]).strip() if row[2] else '',
                    'objectif': str(row[3]).strip() if row[3] else '',
                    'macro_competence': comp,
                    'categorie': str(row[5]).strip() if row[5] else ''
                })

    wb.close()
    return competences


def load_savoirs(filepath: Path) -> list:
    """Charge les savoirs depuis le fichier XLSX."""
    wb = openpyxl.load_workbook(filepath)

    # Chercher la feuille de données
    sheet_name = None
    for name in wb.sheetnames:
        if 'savoir' in name.lower() and 'finition' not in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    savoirs = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:  # Savoir
            savoirs.append({
                'categorie': str(row[0]).strip() if row[0] else '',
                'sous_categorie': str(row[1]).strip() if row[1] else '',
                'savoir': str(row[2]).strip(),
                'code_ogr': str(row[3]).strip() if row[3] else ''
            })

    wb.close()
    return savoirs


# ===========================================================================
# Import en base de données
# ===========================================================================

def import_fiches(repo: Repository, rome_data: dict, verbose: bool = True):
    """Importe les fiches ROME dans la base de données."""
    fiches = rome_data['fiches']
    total = len(fiches)
    imported = 0
    updated = 0
    errors = 0

    if verbose:
        print(f"\n{'='*60}")
        print(f"  Import de {total} fiches métiers ROME")
        print(f"{'='*60}\n")

    for i, (code_rome, data) in enumerate(fiches.items(), 1):
        try:
            # Construire les appellations comme liste de noms
            appellations = [a['nom'] for a in data['appellations']]

            fiche = FicheMetier(
                id=code_rome,
                code_rome=code_rome,
                code_ogr=data['code_ogr'],
                nom_masculin=data['nom_masculin'],
                nom_feminin=data['nom_feminin'],
                nom_epicene=data['nom_epicene'],
                description=f"Fiche métier ROME {code_rome} - {data['nom_complet']}",
                description_courte=data['nom_complet'],
                competences=[],
                competences_transversales=[],
                formations=[],
                certifications=[],
                conditions_travail=[],
                environnements=[],
                metiers_proches=[],
                secteurs_activite=[data['domaine']],
                salaires=SalairesMetier(),
                perspectives=PerspectivesMetier(
                    tendance=TendanceMetier.STABLE
                ),
                metadata=MetadataFiche(
                    statut=StatutFiche.BROUILLON,
                    source="ROME data.gouv.fr (sept. 2025)",
                    tags=[
                        f"domaine:{data['domaine_code']}",
                        f"sous-domaine:{data['sous_domaine_code']}",
                        data['sous_domaine']
                    ] + ([f"appellations:{len(appellations)}"] if appellations else []),
                    auteur="import_rome.py"
                )
            )

            # Vérifier si la fiche existe déjà
            existing = repo.get_fiche(code_rome)
            if existing:
                repo.update_fiche(fiche)
                updated += 1
            else:
                repo.create_fiche(fiche)
                imported += 1

            if verbose and i % 100 == 0:
                print(f"  [{i:4d}/{total}] {code_rome} - {data['nom_masculin'][:50]}")

        except Exception as e:
            errors += 1
            if verbose:
                print(f"  ERREUR {code_rome}: {e}")

    if verbose:
        print(f"\n{'='*60}")
        print(f"  Resultat de l'import")
        print(f"{'='*60}")
        print(f"  Nouvelles fiches : {imported}")
        print(f"  Fiches mises a jour : {updated}")
        print(f"  Erreurs : {errors}")
        print(f"  Total en base : {repo.count_fiches()}")
        print(f"{'='*60}\n")

    return {'imported': imported, 'updated': updated, 'errors': errors}


def print_domaines(rome_data: dict):
    """Affiche les domaines et sous-domaines importés."""
    print(f"\n  14 grands domaines ROME :")
    for code, nom in sorted(rome_data['domaines'].items()):
        nb_fiches = sum(
            1 for f in rome_data['fiches'].values()
            if f['domaine_code'] == code
        )
        print(f"    {code} - {nom} ({nb_fiches} fiches)")


# ===========================================================================
# Main
# ===========================================================================

def main():
    """Point d'entrée du script d'import ROME."""
    print("\n" + "="*60)
    print("  IMPORT REFERENTIEL ROME")
    print("  Source : data.gouv.fr (sept. 2025)")
    print("="*60)

    config = get_config()
    data_dir = config.base_path / "data" / "rome"

    # Vérifier les fichiers
    arbo_file = data_dir / "arborescence_principale.xlsx"
    comp_file = data_dir / "arborescence_competences.xlsx"
    savoirs_file = data_dir / "arborescence_savoirs.xlsx"

    if not arbo_file.exists():
        print(f"\n  ERREUR: Fichier non trouve: {arbo_file}")
        print("  Telecharger depuis: https://www.data.gouv.fr/datasets/repertoire-operationnel-des-metiers-et-des-emplois-rome")
        sys.exit(1)

    # 1. Charger l'arborescence principale
    print("\n  [1/3] Chargement de l'arborescence principale...")
    rome_data = load_arborescence_principale(arbo_file)
    print(f"        {len(rome_data['domaines'])} domaines")
    print(f"        {len(rome_data['sous_domaines'])} sous-domaines")
    print(f"        {len(rome_data['fiches'])} fiches metiers")
    total_app = sum(len(f['appellations']) for f in rome_data['fiches'].values())
    print(f"        {total_app} appellations")

    # 2. Charger les compétences
    if comp_file.exists():
        print("\n  [2/3] Chargement des competences...")
        competences = load_competences(comp_file)
        print(f"        {len(competences)} macro-competences")
    else:
        print("\n  [2/3] Fichier competences non trouve, ignore.")
        competences = []

    # 3. Charger les savoirs
    if savoirs_file.exists():
        print("\n  [3/3] Chargement des savoirs...")
        savoirs = load_savoirs(savoirs_file)
        print(f"        {len(savoirs)} savoirs")
    else:
        print("\n  [3/3] Fichier savoirs non trouve, ignore.")
        savoirs = []

    # Afficher les domaines
    print_domaines(rome_data)

    # Import en base
    print("\n  Initialisation de la base de donnees...")
    repo = Repository(config.db_path)
    repo.init_db()

    result = import_fiches(repo, rome_data)

    # Résumé final
    print("  Commandes utiles :")
    print("    python main.py list                  # Lister les fiches")
    print("    python main.py stats                 # Statistiques")
    print("    python main.py show A1101            # Voir une fiche")
    print("    python main.py search \"developpeur\"  # Rechercher")
    print()


if __name__ == "__main__":
    main()
