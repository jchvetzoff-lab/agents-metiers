#!/usr/bin/env python3
"""
Import des fiches ROME vers l'API backend (PostgreSQL sur Render).
Lit les fichiers XLSX locaux et envoie les fiches via POST /api/fiches.
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
import requests
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

# Configuration
API_URL = "https://agents-metiers.onrender.com"
BATCH_SIZE = 50  # Nombre de fiches par batch
MAX_WORKERS = 5  # Requêtes parallèles


def parse_nom_genre(nom_complet: str) -> dict:
    """Parse un nom au format 'Masculin / Féminin complément' du ROME."""
    if ' / ' not in nom_complet:
        return {
            'masculin': nom_complet,
            'feminin': nom_complet,
            'epicene': nom_complet
        }

    parts = nom_complet.split(' / ', 1)
    masculin_part = parts[0].strip()
    feminin_rest = parts[1].strip()

    masc_words = masculin_part.split()
    fem_words = feminin_rest.split()
    nb_mots_nom = len(masc_words)

    if len(fem_words) > nb_mots_nom:
        fem_nom = ' '.join(fem_words[:nb_mots_nom])
        complement = ' '.join(fem_words[nb_mots_nom:])
        masculin = f"{masculin_part} {complement}"
        feminin = f"{fem_nom} {complement}"
    else:
        masculin = masculin_part
        feminin = feminin_rest

    epicene = nom_complet

    return {
        'masculin': masculin,
        'feminin': feminin,
        'epicene': epicene
    }


def load_arborescence_principale(filepath: Path) -> dict:
    """Charge l'arborescence principale du ROME."""
    print(f"  Lecture de {filepath.name}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[1]]

    domaines = {}
    sous_domaines = {}
    fiches = {}

    current_domaine = ""
    current_domaine_nom = ""
    current_sous_domaine = ""
    current_sous_domaine_nom = ""

    for row in ws.iter_rows(min_row=2, values_only=True):
        col1 = str(row[0]).strip() if row[0] else ""
        col2 = str(row[1]).strip() if row[1] else ""
        col3 = str(row[2]).strip() if row[2] else ""
        col4 = str(row[3]).strip() if row[3] else ""
        col5 = str(row[4]).strip() if row[4] else ""

        if col1 and not col2 and not col3 and col4:
            current_domaine = col1
            current_domaine_nom = col4
            domaines[col1] = col4
            continue

        if col1 and col2 and not col3 and col4:
            current_sous_domaine = f"{col1}{col2}"
            current_sous_domaine_nom = col4
            sous_domaines[current_sous_domaine] = col4
            continue

        if col3:
            code_rome = f"{col1}{col2}{col3}"

            if code_rome not in fiches:
                noms = parse_nom_genre(col4)
                fiches[code_rome] = {
                    'nom_complet': col4,
                    'nom_masculin': noms['masculin'],
                    'nom_feminin': noms['feminin'],
                    'nom_epicene': noms['epicene'],
                    'code_ogr': col5,
                    'domaine': current_domaine_nom,
                    'domaine_code': current_domaine,
                    'sous_domaine': current_sous_domaine_nom,
                    'sous_domaine_code': current_sous_domaine,
                    'appellations': []
                }
            else:
                fiches[code_rome]['appellations'].append({
                    'nom': col4,
                    'code_ogr': col5
                })

    wb.close()
    return {
        'domaines': domaines,
        'sous_domaines': sous_domaines,
        'fiches': fiches
    }


def create_fiche_payload(code_rome: str, data: dict) -> dict:
    """Crée le payload JSON pour l'API."""
    return {
        "code_rome": code_rome,
        "nom_masculin": data['nom_masculin'],
        "nom_feminin": data['nom_feminin'],
        "nom_epicene": data['nom_epicene'],
        "description": f"Fiche métier ROME {code_rome} - {data['nom_complet']}",
        "secteurs_activite": [data['domaine']]
    }


def send_fiche(session: requests.Session, code_rome: str, payload: dict) -> tuple:
    """Envoie une fiche à l'API. Retourne (code_rome, success, message)."""
    try:
        response = session.post(
            f"{API_URL}/api/fiches",
            json=payload,
            timeout=30
        )
        if response.status_code == 200:
            return (code_rome, True, "created")
        elif response.status_code == 400 and "existe" in response.text.lower():
            return (code_rome, True, "exists")
        else:
            return (code_rome, False, f"HTTP {response.status_code}: {response.text[:100]}")
    except Exception as e:
        return (code_rome, False, str(e))


def import_fiches_parallel(fiches: dict) -> dict:
    """Importe les fiches en parallèle."""
    total = len(fiches)
    created = 0
    existed = 0
    errors = 0
    error_details = []

    print(f"\n  Import de {total} fiches vers {API_URL}")
    print(f"  Workers: {MAX_WORKERS}, Batch: {BATCH_SIZE}")
    print("  " + "=" * 50)

    session = requests.Session()

    # Préparer tous les payloads
    payloads = [
        (code, create_fiche_payload(code, data))
        for code, data in fiches.items()
    ]

    start_time = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(send_fiche, session, code, payload): code
            for code, payload in payloads
        }

        for i, future in enumerate(as_completed(futures), 1):
            code_rome, success, message = future.result()

            if success:
                if message == "created":
                    created += 1
                else:
                    existed += 1
            else:
                errors += 1
                error_details.append(f"{code_rome}: {message}")

            # Progress
            if i % 100 == 0 or i == total:
                elapsed = time.time() - start_time
                rate = i / elapsed if elapsed > 0 else 0
                eta = (total - i) / rate if rate > 0 else 0
                print(f"  [{i:4d}/{total}] Created: {created}, Existed: {existed}, Errors: {errors} | {rate:.1f}/s | ETA: {eta:.0f}s")

    elapsed = time.time() - start_time

    print("  " + "=" * 50)
    print(f"  Terminé en {elapsed:.1f}s ({total/elapsed:.1f} fiches/s)")
    print(f"  Créées: {created}")
    print(f"  Existantes: {existed}")
    print(f"  Erreurs: {errors}")

    if error_details and errors <= 10:
        print("\n  Détail des erreurs:")
        for err in error_details[:10]:
            print(f"    - {err}")

    return {
        'created': created,
        'existed': existed,
        'errors': errors,
        'total': total,
        'elapsed': elapsed
    }


def main():
    print("\n" + "=" * 60)
    print("  IMPORT ROME vers API (PostgreSQL Render)")
    print("=" * 60)

    # Chemin des données
    data_dir = Path(__file__).parent.parent / "data" / "rome"
    arbo_file = data_dir / "arborescence_principale.xlsx"

    if not arbo_file.exists():
        print(f"\n  ERREUR: Fichier non trouvé: {arbo_file}")
        sys.exit(1)

    # Test connexion API
    print("\n  [1/3] Test connexion API...")
    try:
        r = requests.get(f"{API_URL}/", timeout=10)
        if r.status_code == 200:
            print(f"        API OK: {API_URL}")
        else:
            print(f"        ERREUR: HTTP {r.status_code}")
            sys.exit(1)
    except Exception as e:
        print(f"        ERREUR connexion: {e}")
        sys.exit(1)

    # Stats avant
    print("\n  [2/3] Stats avant import...")
    try:
        r = requests.get(f"{API_URL}/api/stats", timeout=10)
        stats = r.json()
        print(f"        Fiches en base: {stats['total']}")
    except:
        print("        Impossible de récupérer les stats")

    # Charger les données
    print("\n  [3/3] Chargement des données ROME...")
    rome_data = load_arborescence_principale(arbo_file)
    print(f"        {len(rome_data['domaines'])} domaines")
    print(f"        {len(rome_data['fiches'])} fiches")

    # Confirmation (skip with --yes flag)
    if "--yes" not in sys.argv:
        print(f"\n  Prêt à importer {len(rome_data['fiches'])} fiches.")
        confirm = input("  Continuer ? (o/N) : ").strip().lower()
        if confirm != 'o':
            print("  Annulé.")
            sys.exit(0)
    else:
        print(f"\n  Import de {len(rome_data['fiches'])} fiches (--yes flag)")

    # Import
    result = import_fiches_parallel(rome_data['fiches'])

    # Stats après
    print("\n  Stats après import:")
    try:
        r = requests.get(f"{API_URL}/api/stats", timeout=10)
        stats = r.json()
        print(f"        Fiches en base: {stats['total']}")
        print(f"        - Brouillons: {stats['brouillons']}")
        print(f"        - En validation: {stats['en_validation']}")
        print(f"        - Publiées: {stats['publiees']}")
    except:
        pass

    print("\n" + "=" * 60)
    print("  IMPORT TERMINÉ")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
