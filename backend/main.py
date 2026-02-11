"""
Backend FastAPI pour Agents Métiers Web.
Expose une API REST pour accéder à la base de données SQLite et aux agents IA.
"""
from fastapi import FastAPI, HTTPException, Query, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import sys
import os
import json
import re
from pathlib import Path
from sqlalchemy import text as sa_text

# Ajouter le chemin vers le projet agents-metiers existant
AGENTS_METIERS_PATH = Path(__file__).parent.parent.parent / "agents-metiers"
sys.path.insert(0, str(AGENTS_METIERS_PATH))

from database.repository import Repository
from database.models import (
    StatutFiche, FicheMetier, VarianteFiche, MobiliteMetier, TypesContrats,
    AuditLog, TypeEvenement, GenreGrammatical, TrancheAge, FormatContenu, LangueSupporte,
)
from config import get_config

app = FastAPI(
    title="Agents Métiers API",
    description="API REST pour la gestion des fiches métiers avec IA",
    version="1.0.0"
)

# Configuration CORS pour le frontend Next.js
# ALLOWED_ORIGINS env var: comma-separated list, or "*" (default) for all origins
# JWT auth is header-based (not cookies), so allow_credentials=False is safe with "*"
_allowed_origins = os.environ.get("ALLOWED_ORIGINS", "*")
if _allowed_origins == "*":
    _cors_origins = ["*"]
    _cors_regex = None
    _cors_credentials = False
else:
    _cors_origins = [o.strip() for o in _allowed_origins.split(",")]
    _cors_regex = r"https://.*\.netlify\.app"
    _cors_credentials = True

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_origin_regex=_cors_regex,
    allow_credentials=_cors_credentials,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialisation du repository
config = get_config()
# Utilise PostgreSQL si DATABASE_URL est défini (production), sinon SQLite (dev)
repo = Repository(
    db_path=config.db_path if not config.database.database_url else None,
    database_url=config.database.database_url
)
repo.init_db()

# Migration : ajouter les nouvelles colonnes si elles n'existent pas
def run_migrations():
    """Ajoute les colonnes manquantes à la table fiches_metiers."""
    from sqlalchemy import text, inspect
    engine = repo.engine
    inspector = inspect(engine)
    existing_cols = {c["name"] for c in inspector.get_columns("fiches_metiers")}
    new_columns = {
        "savoirs": "JSON DEFAULT '[]'",
        "acces_metier": "TEXT",
        "missions_principales": "JSON DEFAULT '[]'",
        "types_contrats": "JSON DEFAULT '{}'",
        "mobilite": "JSON DEFAULT '{}'",
    }
    with engine.begin() as conn:
        for col_name, col_type in new_columns.items():
            if col_name not in existing_cols:
                conn.execute(text(f'ALTER TABLE fiches_metiers ADD COLUMN {col_name} {col_type}'))

try:
    run_migrations()
except Exception as e:
    print(f"Migration warning: {e}")

# Migration auth : créer la table users
from backend.auth import (
    create_users_table, UserCreate, UserLogin, UserResponse, UserDB,
    hash_password, verify_password, create_token, decode_token, get_current_user,
    JWT_SECRET_KEY
)
try:
    create_users_table(repo.engine)
except Exception as e:
    print(f"Users table migration warning: {e}")

# Migration : créer la table offres_cache
def create_offres_cache_table(engine):
    """Crée la table offres_cache si elle n'existe pas."""
    from sqlalchemy import inspect as sa_inspect
    inspector = sa_inspect(engine)
    if "offres_cache" not in inspector.get_table_names():
        from sqlalchemy import text
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE offres_cache (
                    id SERIAL PRIMARY KEY,
                    code_rome VARCHAR(10) NOT NULL,
                    region VARCHAR(5),
                    offre_id VARCHAR(50) NOT NULL,
                    titre VARCHAR(500),
                    entreprise VARCHAR(300),
                    lieu VARCHAR(300),
                    type_contrat VARCHAR(50),
                    salaire VARCHAR(200),
                    experience VARCHAR(200),
                    date_publication TIMESTAMP,
                    url VARCHAR(700),
                    fetched_at TIMESTAMP NOT NULL DEFAULT NOW()
                )
            """))
            conn.execute(text("CREATE INDEX idx_offres_cache_rome ON offres_cache (code_rome)"))
            conn.execute(text("CREATE INDEX idx_offres_cache_fetched ON offres_cache (fetched_at)"))
        print("Created offres_cache table")

try:
    create_offres_cache_table(repo.engine)
except Exception as e:
    print(f"Offres cache table migration warning: {e}")


# ==================== HELPERS ====================

CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-20250514")
ADMIN_KEY = os.environ.get("ADMIN_KEY", JWT_SECRET_KEY)  # Separate admin key; falls back to JWT secret if not set

import time as _time

def claude_call_with_retry(client, *, model, max_tokens, messages, max_retries=5):
    """Call Claude API with retry + exponential backoff on 429 rate-limit errors."""
    import anthropic
    for attempt in range(max_retries):
        try:
            return client.messages.create(model=model, max_tokens=max_tokens, messages=messages)
        except anthropic.RateLimitError as e:
            if attempt == max_retries - 1:
                raise
            wait = min(2 ** attempt * 15, 120)  # 15s, 30s, 60s, 120s
            print(f"Rate limit hit, waiting {wait}s before retry {attempt + 2}/{max_retries}...")
            _time.sleep(wait)


def log_action(type_evt: TypeEvenement, description: str, code_rome: str = None, agent: str = "Frontend", validateur: str = None):
    """Enregistre un audit log."""
    try:
        log = AuditLog(
            type_evenement=type_evt,
            agent=agent,
            code_rome=code_rome,
            description=description,
            validateur=validateur,
        )
        repo.add_audit_log(log)
    except Exception as e:
        print(f"Audit log error: {e}")


# ==================== MODELS ====================

class FicheMetierResponse(BaseModel):
    """Modèle de réponse pour une fiche métier."""
    code_rome: str
    nom_masculin: str
    nom_feminin: str
    nom_epicene: str
    statut: str
    description: Optional[str] = None
    description_courte: Optional[str] = None
    date_creation: datetime
    date_maj: datetime
    version: int
    # Données enrichies
    has_competences: bool = False
    has_formations: bool = False
    has_salaires: bool = False
    has_perspectives: bool = False
    nb_variantes: int = 0


class StatsResponse(BaseModel):
    """Statistiques globales."""
    total: int
    brouillons: int
    en_validation: int
    publiees: int
    archivees: int


# ==================== ROUTES ====================

@app.get("/")
async def root():
    """Page d'accueil de l'API."""
    return {
        "message": "Agents Métiers API",
        "version": "1.0.0",
        "endpoints": {
            "stats": "/api/stats",
            "fiches": "/api/fiches",
            "fiche_detail": "/api/fiches/{code_rome}",
            "variantes": "/api/fiches/{code_rome}/variantes"
        }
    }


@app.get("/api/stats", response_model=StatsResponse)
async def get_stats():
    """Récupère les statistiques globales."""
    try:
        total = repo.count_fiches()
        brouillons = repo.count_fiches(StatutFiche.BROUILLON)
        en_validation = repo.count_fiches(StatutFiche.EN_VALIDATION)
        publiees = repo.count_fiches(StatutFiche.PUBLIEE)
        archivees = repo.count_fiches(StatutFiche.ARCHIVEE)

        return StatsResponse(
            total=total,
            brouillons=brouillons,
            en_validation=en_validation,
            publiees=publiees,
            archivees=archivees
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/fiches")
async def get_fiches(
    statut: Optional[str] = Query(None, description="Filtrer par statut"),
    search: Optional[str] = Query(None, description="Recherche textuelle"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0)
):
    """Liste les fiches métiers avec filtres et pagination."""
    try:
        from sqlalchemy import select, func, or_
        from database.models import FicheMetierDB

        # Convertir le statut si fourni
        statut_enum = None
        if statut:
            try:
                statut_enum = StatutFiche(statut.lower())
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Statut invalide: {statut}")

        # Construire la requête SQL directement
        with repo.session() as session:
            query = select(FicheMetierDB)
            count_query = select(func.count(FicheMetierDB.id))

            # Filtre statut
            if statut_enum:
                query = query.where(FicheMetierDB.statut == statut_enum.value)
                count_query = count_query.where(FicheMetierDB.statut == statut_enum.value)

            # Filtre recherche en SQL (ILIKE sur code_rome, nom_masculin, nom_feminin, nom_epicene)
            if search:
                search_pattern = f"%{search}%"
                search_filter = or_(
                    FicheMetierDB.code_rome.ilike(search_pattern),
                    FicheMetierDB.nom_masculin.ilike(search_pattern),
                    FicheMetierDB.nom_feminin.ilike(search_pattern),
                    FicheMetierDB.nom_epicene.ilike(search_pattern),
                )
                query = query.where(search_filter)
                count_query = count_query.where(search_filter)

            # Compter le total AVANT pagination
            total = session.execute(count_query).scalar()

            # Appliquer pagination + tri par date de modification decroissant
            query = query.order_by(FicheMetierDB.date_maj.desc()).limit(limit).offset(offset)
            db_fiches = session.execute(query).scalars().all()

            # Convertir en réponse
            results = []
            for db_fiche in db_fiches:
                fiche = db_fiche.to_pydantic()
                nb_variantes = repo.count_variantes(fiche.code_rome)
                results.append(FicheMetierResponse(
                    code_rome=fiche.code_rome,
                    nom_masculin=fiche.nom_masculin,
                    nom_feminin=fiche.nom_feminin,
                    nom_epicene=fiche.nom_epicene,
                    statut=fiche.metadata.statut.value,
                    description=fiche.description,
                    description_courte=fiche.description_courte,
                    date_creation=fiche.metadata.date_creation,
                    date_maj=fiche.metadata.date_maj,
                    version=fiche.metadata.version,
                    has_competences=bool(fiche.competences or fiche.competences_transversales),
                    has_formations=bool(fiche.formations),
                    has_salaires=bool(fiche.salaires),
                    has_perspectives=bool(fiche.perspectives),
                    nb_variantes=nb_variantes
                ))

        return {
            "total": total,
            "limit": limit,
            "offset": offset,
            "results": results
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class FicheMetierCreate(BaseModel):
    """Modèle pour créer une fiche métier."""
    code_rome: str
    nom_masculin: str
    nom_feminin: str
    nom_epicene: str
    definition: Optional[str] = None
    description: Optional[str] = None


@app.post("/api/fiches", status_code=201)
async def create_fiche(fiche_data: FicheMetierCreate, current_user: dict = Depends(get_current_user)):
    """Crée une nouvelle fiche métier."""
    try:
        # Vérifier si la fiche existe déjà
        existing = repo.get_fiche(fiche_data.code_rome)
        if existing:
            raise HTTPException(status_code=400, detail=f"La fiche {fiche_data.code_rome} existe déjà")

        # Créer la fiche
        from database.models import FicheMetier, MetadataFiche, StatutFiche

        nouvelle_fiche = FicheMetier(
            id=fiche_data.code_rome,  # id = code_rome
            code_rome=fiche_data.code_rome,
            nom_masculin=fiche_data.nom_masculin,
            nom_feminin=fiche_data.nom_feminin,
            nom_epicene=fiche_data.nom_epicene,
            description=fiche_data.definition or fiche_data.description or "",
            metadata=MetadataFiche(
                statut=StatutFiche.BROUILLON,
                version=1
            )
        )

        fiche_creee = repo.create_fiche(nouvelle_fiche)

        user_name = current_user.get("name", "Utilisateur")
        log_action(
            TypeEvenement.CREATION,
            f"Création de la fiche {fiche_data.code_rome} ({fiche_data.nom_epicene}) par {user_name}",
            code_rome=fiche_data.code_rome,
            agent=user_name,
        )

        return {
            "message": "Fiche créée avec succès",
            "code_rome": fiche_creee.code_rome,
            "nom_masculin": fiche_creee.nom_masculin,
            "statut": fiche_creee.metadata.statut.value
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la création: {str(e)}")


@app.get("/api/fiches/{code_rome}")
async def get_fiche_detail(code_rome: str):
    """Récupère le détail complet d'une fiche métier."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        return {
            "code_rome": fiche.code_rome,
            "nom_masculin": fiche.nom_masculin,
            "nom_feminin": fiche.nom_feminin,
            "nom_epicene": fiche.nom_epicene,
            "statut": fiche.metadata.statut.value,
            "description": fiche.description,
            "description_courte": fiche.description_courte,
            "missions_principales": fiche.missions_principales,
            "acces_metier": fiche.acces_metier,
            "competences": fiche.competences,
            "competences_transversales": fiche.competences_transversales,
            "savoirs": fiche.savoirs,
            "formations": fiche.formations,
            "certifications": fiche.certifications,
            "conditions_travail": fiche.conditions_travail,
            "environnements": fiche.environnements,
            "salaires": fiche.salaires.model_dump() if fiche.salaires else None,
            "perspectives": fiche.perspectives.model_dump() if fiche.perspectives else None,
            "types_contrats": fiche.types_contrats.model_dump() if fiche.types_contrats else None,
            "mobilite": fiche.mobilite.model_dump() if fiche.mobilite else None,
            "secteurs_activite": fiche.secteurs_activite,
            "date_creation": fiche.metadata.date_creation,
            "date_maj": fiche.metadata.date_maj,
            "version": fiche.metadata.version,
            "nb_variantes": repo.count_variantes(code_rome)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class FicheMetierUpdate(BaseModel):
    """Modèle pour mettre à jour une fiche métier."""
    description: Optional[str] = None
    description_courte: Optional[str] = None
    missions_principales: Optional[List[str]] = None
    acces_metier: Optional[str] = None
    competences: Optional[List[str]] = None
    competences_transversales: Optional[List[str]] = None
    savoirs: Optional[List[str]] = None
    formations: Optional[List[str]] = None
    certifications: Optional[List[str]] = None
    conditions_travail: Optional[List[str]] = None
    environnements: Optional[List[str]] = None
    secteurs_activite: Optional[List[str]] = None
    salaires: Optional[dict] = None
    perspectives: Optional[dict] = None
    types_contrats: Optional[dict] = None
    mobilite: Optional[dict] = None
    statut: Optional[str] = None


@app.patch("/api/fiches/{code_rome}")
async def update_fiche(code_rome: str, update_data: FicheMetierUpdate, user=Depends(get_current_user)):
    """Met à jour une fiche métier existante."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        # Appliquer les mises à jour
        fiche_dict = fiche.model_dump()
        update_dict = update_data.model_dump(exclude_none=True)

        for key, value in update_dict.items():
            if key == "statut":
                fiche_dict["metadata"]["statut"] = value
            elif key in ("salaires", "perspectives", "types_contrats", "mobilite") and value:
                fiche_dict[key] = value
            else:
                fiche_dict[key] = value

        # Mettre à jour les métadonnées
        fiche_dict["metadata"]["date_maj"] = datetime.now()
        fiche_dict["metadata"]["version"] = fiche_dict["metadata"].get("version", 1) + 1

        # Recréer la fiche et sauvegarder
        updated_fiche = FicheMetier(**fiche_dict)
        repo.update_fiche(updated_fiche)

        return {
            "message": "Fiche mise à jour",
            "code_rome": code_rome,
            "version": updated_fiche.metadata.version
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur mise à jour: {str(e)}")


@app.get("/api/fiches/{code_rome}/variantes")
async def get_variantes(code_rome: str):
    """Récupère toutes les variantes d'une fiche."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        variantes = repo.get_all_variantes(code_rome)

        return {
            "code_rome": code_rome,
            "total_variantes": len(variantes),
            "variantes": [
                {
                    "id": v.id,
                    "langue": v.langue.value,
                    "tranche_age": v.tranche_age.value,
                    "format_contenu": v.format_contenu.value,
                    "genre": v.genre.value,
                    "nom": v.nom,
                    "description_courte": v.description_courte,
                    "date_maj": v.date_maj
                }
                for v in variantes
            ]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/fiches/{code_rome}/variantes/{variante_id}")
async def get_variante_detail(code_rome: str, variante_id: int):
    """Récupère le détail d'une variante spécifique."""
    try:
        # TODO: Implémenter get_variante_by_id dans le repository
        variantes = repo.get_all_variantes(code_rome)
        variante = next((v for v in variantes if v.id == variante_id), None)

        if not variante:
            raise HTTPException(status_code=404, detail=f"Variante {variante_id} non trouvée")

        return {
            "id": variante.id,
            "code_rome": variante.code_rome,
            "langue": variante.langue.value,
            "tranche_age": variante.tranche_age.value,
            "format_contenu": variante.format_contenu.value,
            "genre": variante.genre.value,
            "nom": variante.nom,
            "description": variante.description,
            "description_courte": variante.description_courte,
            "competences": variante.competences,
            "competences_transversales": variante.competences_transversales,
            "formations": variante.formations,
            "certifications": variante.certifications,
            "conditions_travail": variante.conditions_travail,
            "environnements": variante.environnements,
            "date_creation": variante.date_creation,
            "date_maj": variante.date_maj,
            "version": variante.version
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/audit-logs")
async def get_audit_logs(limit: int = Query(15, ge=1, le=100)):
    """Récupère les logs d'audit."""
    try:
        logs = repo.get_audit_logs(limit=limit)

        return {
            "total": len(logs),
            "logs": [
                {
                    "id": log.id,
                    "type_evenement": log.type_evenement.value,
                    "description": log.description,
                    "code_rome": log.code_rome,
                    "agent": log.agent,
                    "validateur": log.validateur,
                    "timestamp": log.timestamp
                }
                for log in logs
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ACTIONS IA ====================

ENRICH_PROMPT = """Tu es un expert en ressources humaines et en rédaction de fiches métiers en France.
Génère le contenu COMPLET et EXHAUSTIF pour la fiche métier suivante. Chaque section DOIT être remplie avec le MAXIMUM d'éléments pertinents.

Métier : {nom_masculin} / {nom_feminin}
Code ROME : {code_rome}
{contexte}

Réponds UNIQUEMENT avec un objet JSON valide (sans texte avant ou après) contenant :
{{
    "description": "Description complète et détaillée du métier en 5-8 phrases. Couvrir le rôle, les responsabilités, le contexte, l'importance du métier.",
    "description_courte": "Description en 1 phrase percutante (max 200 caractères).",
    "missions_principales": ["8 à 12 missions principales du métier - être exhaustif"],
    "acces_metier": "Texte détaillé décrivant comment accéder à ce métier : formations initiales, parcours alternatifs, expérience requise, prérequis, VAE possible. 4-6 phrases.",
    "competences": ["10 à 15 compétences techniques clés (savoir-faire) - être très complet"],
    "competences_transversales": ["6 à 10 soft skills / savoir-être essentiels"],
    "savoirs": ["8 à 12 connaissances théoriques nécessaires"],
    "formations": ["5 à 8 formations ou diplômes typiques (du CAP au Master si applicable)"],
    "certifications": ["3 à 6 certifications professionnelles reconnues dans le domaine"],
    "conditions_travail": ["5 à 8 conditions de travail caractéristiques (horaires, lieu, contraintes, avantages)"],
    "environnements": ["4 à 6 types de structures/environnements de travail"],
    "secteurs_activite": ["4 à 6 secteurs d'activité"],
    "salaires": {{
        "junior": {{"min": 25000, "max": 35000, "median": 30000}},
        "confirme": {{"min": 35000, "max": 50000, "median": 42000}},
        "senior": {{"min": 50000, "max": 70000, "median": 58000}}
    }},
    "perspectives": {{
        "tension": 0.6,
        "tendance": "stable",
        "evolution_5ans": "Analyse courte de l'évolution du métier dans les 5 prochaines années.",
        "nombre_offres": 5000,
        "taux_insertion": 0.75
    }},
    "types_contrats": {{
        "cdi": 60,
        "cdd": 20,
        "interim": 15,
        "autre": 5
    }},
    "mobilite": {{
        "metiers_proches": [
            {{"nom": "Métier proche 1", "contexte": "Compétences communes en..."}},
            {{"nom": "Métier proche 2", "contexte": "Compétences communes en..."}}
        ],
        "evolutions": [
            {{"nom": "Évolution possible 1", "contexte": "Après expérience en..."}},
            {{"nom": "Évolution possible 2", "contexte": "Avec formation complémentaire en..."}}
        ]
    }}
}}

Notes :
- Les salaires sont en euros brut annuel pour la France en 2025.
- "tension" est un float entre 0 (peu de demande) et 1 (très forte demande).
- "tendance" est "emergence", "stable" ou "disparition".
- "nombre_offres" est une estimation du nombre d'offres d'emploi par an en France.
- "taux_insertion" est un float entre 0 et 1.
- "types_contrats" : les pourcentages doivent totaliser 100.
- Sois factuel et précis. Pas de formulations vagues."""


@app.post("/api/fiches/{code_rome}/enrich")
def enrich_fiche(code_rome: str, current_user: dict = Depends(get_current_user)):
    """Enrichit une fiche métier avec Claude API."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        api_key = os.getenv("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY non configurée")

        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        contexte = ""
        if fiche.description:
            contexte = f"\nDescription existante : {fiche.description}"

        prompt = ENRICH_PROMPT.format(
            nom_masculin=fiche.nom_masculin,
            nom_feminin=fiche.nom_feminin,
            code_rome=fiche.code_rome,
            contexte=contexte,
        )

        response = claude_call_with_retry(
            client,
            model=CLAUDE_MODEL,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )

        content = response.content[0].text.strip()
        json_match = re.search(r'\{[\s\S]*\}', content)
        if not json_match:
            raise HTTPException(status_code=500, detail="Réponse Claude invalide (pas de JSON)")

        data = json.loads(json_match.group())

        # Build update dict
        fiche_dict = fiche.model_dump()
        for key in [
            "description", "description_courte", "missions_principales",
            "acces_metier", "competences", "competences_transversales",
            "savoirs", "formations", "certifications", "conditions_travail",
            "environnements", "secteurs_activite",
        ]:
            if key in data:
                fiche_dict[key] = data[key]

        if "salaires" in data:
            fiche_dict["salaires"] = data["salaires"]
        if "perspectives" in data:
            fiche_dict["perspectives"] = data["perspectives"]
        if "types_contrats" in data:
            fiche_dict["types_contrats"] = data["types_contrats"]
        if "mobilite" in data:
            fiche_dict["mobilite"] = data["mobilite"]

        fiche_dict["metadata"]["statut"] = "en_validation"
        fiche_dict["metadata"]["date_maj"] = datetime.now()
        fiche_dict["metadata"]["version"] = fiche_dict["metadata"].get("version", 1) + 1

        updated_fiche = FicheMetier(**fiche_dict)
        repo.update_fiche(updated_fiche)

        user_name = current_user.get("name", "Utilisateur")
        log_action(
            TypeEvenement.MODIFICATION,
            f"Enrichissement IA de {code_rome} ({fiche.nom_epicene}) par {user_name} - v{updated_fiche.metadata.version}",
            code_rome=code_rome,
            agent=user_name,
        )

        return {
            "message": "Fiche enrichie avec succès",
            "code_rome": code_rome,
            "nom": fiche.nom_epicene,
            "version": updated_fiche.metadata.version,
        }

    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Erreur parsing JSON Claude: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur enrichissement: {str(e)}")


@app.post("/api/fiches/{code_rome}/publish")
async def publish_fiche(code_rome: str, current_user: dict = Depends(get_current_user)):
    """Publie une fiche métier (change le statut en publiee)."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        fiche_dict = fiche.model_dump()
        fiche_dict["metadata"]["statut"] = "publiee"
        fiche_dict["metadata"]["date_maj"] = datetime.now()
        fiche_dict["metadata"]["version"] = fiche_dict["metadata"].get("version", 1) + 1

        updated_fiche = FicheMetier(**fiche_dict)
        repo.update_fiche(updated_fiche)

        user_name = current_user.get("name", "Utilisateur")
        log_action(
            TypeEvenement.PUBLICATION,
            f"Publication de {code_rome} ({fiche.nom_epicene}) par {user_name}",
            code_rome=code_rome,
            agent=user_name,
        )

        return {
            "message": "Fiche publiée",
            "code_rome": code_rome,
            "statut": "publiee",
            "publie_par": user_name,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur publication: {str(e)}")


class PublishBatchRequest(BaseModel):
    codes_rome: List[str]


@app.post("/api/fiches/publish-batch")
async def publish_batch(request: PublishBatchRequest, current_user: dict = Depends(get_current_user)):
    """Publie plusieurs fiches en masse."""
    results = []
    for code_rome in request.codes_rome:
        try:
            fiche = repo.get_fiche(code_rome)
            if not fiche:
                results.append({"code_rome": code_rome, "status": "error", "message": "Non trouvée"})
                continue

            fiche_dict = fiche.model_dump()
            fiche_dict["metadata"]["statut"] = "publiee"
            fiche_dict["metadata"]["date_maj"] = datetime.now()
            fiche_dict["metadata"]["version"] = fiche_dict["metadata"].get("version", 1) + 1

            updated_fiche = FicheMetier(**fiche_dict)
            repo.update_fiche(updated_fiche)
            results.append({"code_rome": code_rome, "status": "ok", "message": "Publiée"})
            user_name = current_user.get("name", "Utilisateur")
            log_action(
                TypeEvenement.PUBLICATION,
                f"Publication batch de {code_rome} par {user_name}",
                code_rome=code_rome,
                agent=user_name,
            )
        except Exception as e:
            results.append({"code_rome": code_rome, "status": "error", "message": str(e)})

    ok_count = sum(1 for r in results if r["status"] == "ok")
    return {
        "message": f"{ok_count}/{len(request.codes_rome)} fiches publiées",
        "results": results,
    }


# ==================== VALIDATION ====================

VALIDATE_PROMPT = """Tu es un expert qualité en fiches métiers ROME pour France Travail.
Analyse la fiche métier suivante et fournis un rapport de validation détaillé.

Fiche métier :
- Code ROME : {code_rome}
- Nom : {nom_masculin} / {nom_feminin}
- Description : {description}
- Description courte : {description_courte}
- Missions principales : {missions_principales}
- Accès métier : {acces_metier}
- Compétences techniques : {competences}
- Compétences transversales : {competences_transversales}
- Savoirs : {savoirs}
- Formations : {formations}
- Certifications : {certifications}
- Conditions de travail : {conditions_travail}
- Environnements : {environnements}
- Secteurs d'activité : {secteurs_activite}
- Salaires : {salaires}
- Perspectives : {perspectives}
- Types de contrats : {types_contrats}
- Mobilité : {mobilite}

Évalue la fiche sur les critères suivants :
1. **Complétude** : Toutes les sections sont-elles remplies ? Les listes ont-elles assez d'éléments ?
2. **Exactitude** : Les informations sont-elles factuellement correctes pour la France ?
3. **Cohérence** : Les compétences, formations et salaires sont-ils cohérents entre eux ?
4. **Qualité rédactionnelle** : Orthographe, grammaire, clarté, ton professionnel ?
5. **Pertinence** : Les données correspondent-elles bien au métier décrit ?

Réponds UNIQUEMENT avec un objet JSON valide :
{{
    "score": 85,
    "verdict": "approuvee",
    "resume": "Résumé en 1-2 phrases de la qualité globale.",
    "criteres": {{
        "completude": {{"score": 90, "commentaire": "..."}},
        "exactitude": {{"score": 85, "commentaire": "..."}},
        "coherence": {{"score": 80, "commentaire": "..."}},
        "qualite_redactionnelle": {{"score": 90, "commentaire": "..."}},
        "pertinence": {{"score": 85, "commentaire": "..."}}
    }},
    "problemes": ["Problème 1 à corriger", "Problème 2 à corriger"],
    "suggestions": ["Amélioration suggérée 1", "Amélioration suggérée 2"]
}}

Notes :
- "score" est un entier de 0 à 100 (moyenne des 5 critères).
- "verdict" : "approuvee" (score >= 70), "a_corriger" (score 40-69), "rejetee" (score < 40).
- "problemes" : liste des erreurs factuelles ou lacunes importantes (peut être vide).
- "suggestions" : liste d'améliorations optionnelles (peut être vide).
- Sois exigeant mais juste. Une bonne fiche enrichie devrait obtenir 75-90."""


@app.post("/api/fiches/{code_rome}/validate")
def validate_fiche(code_rome: str, current_user: dict = Depends(get_current_user)):
    """Validation IA : Claude analyse la qualité de la fiche et retourne un rapport."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        if fiche.metadata.statut.value == "brouillon" and not fiche.description:
            raise HTTPException(status_code=400, detail="Fiche non enrichie, impossible de valider un brouillon vide")

        api_key = os.getenv("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY non configurée")

        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        prompt = VALIDATE_PROMPT.format(
            code_rome=fiche.code_rome,
            nom_masculin=fiche.nom_masculin,
            nom_feminin=fiche.nom_feminin,
            description=fiche.description or "Non renseignée",
            description_courte=fiche.description_courte or "Non renseignée",
            missions_principales=json.dumps(fiche.missions_principales or [], ensure_ascii=False),
            acces_metier=fiche.acces_metier or "Non renseigné",
            competences=json.dumps(fiche.competences or [], ensure_ascii=False),
            competences_transversales=json.dumps(fiche.competences_transversales or [], ensure_ascii=False),
            savoirs=json.dumps(fiche.savoirs or [], ensure_ascii=False),
            formations=json.dumps(fiche.formations or [], ensure_ascii=False),
            certifications=json.dumps(fiche.certifications or [], ensure_ascii=False),
            conditions_travail=json.dumps(fiche.conditions_travail or [], ensure_ascii=False),
            environnements=json.dumps(fiche.environnements or [], ensure_ascii=False),
            secteurs_activite=json.dumps(fiche.secteurs_activite or [], ensure_ascii=False),
            salaires=json.dumps(fiche.salaires.model_dump() if fiche.salaires else {}, ensure_ascii=False),
            perspectives=json.dumps(fiche.perspectives.model_dump() if fiche.perspectives else {}, ensure_ascii=False),
            types_contrats=json.dumps(fiche.types_contrats.model_dump() if fiche.types_contrats else {}, ensure_ascii=False),
            mobilite=json.dumps(fiche.mobilite.model_dump() if fiche.mobilite else {}, ensure_ascii=False),
        )

        response = claude_call_with_retry(
            client,
            model=CLAUDE_MODEL,
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )

        content = response.content[0].text.strip()
        json_match = re.search(r'\{[\s\S]*\}', content)
        if not json_match:
            raise HTTPException(status_code=500, detail="Réponse Claude invalide (pas de JSON)")

        rapport = json.loads(json_match.group())

        user_name = current_user.get("name", "Utilisateur")
        log_action(
            TypeEvenement.VALIDATION,
            f"Validation IA de {code_rome} ({fiche.nom_epicene}) par {user_name} - Score: {rapport.get('score', '?')}/100 - Verdict: {rapport.get('verdict', '?')}",
            code_rome=code_rome,
            agent=user_name,
        )

        return {
            "message": "Validation IA terminée",
            "code_rome": code_rome,
            "nom": fiche.nom_epicene,
            "rapport": rapport,
        }

    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Erreur parsing JSON Claude: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur validation IA: {str(e)}")


class HumanReviewRequest(BaseModel):
    decision: str  # "approuvee", "a_corriger", "rejetee"
    commentaire: Optional[str] = None


@app.post("/api/fiches/{code_rome}/review")
async def review_fiche(code_rome: str, review: HumanReviewRequest, current_user: dict = Depends(get_current_user)):
    """Validation humaine : approuver, demander des corrections, ou rejeter une fiche."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        valid_decisions = ["approuvee", "a_corriger", "rejetee"]
        if review.decision not in valid_decisions:
            raise HTTPException(
                status_code=400,
                detail=f"Décision invalide. Valeurs acceptées: {valid_decisions}"
            )

        fiche_dict = fiche.model_dump()

        if review.decision == "approuvee":
            fiche_dict["metadata"]["statut"] = "publiee"
        elif review.decision == "a_corriger":
            fiche_dict["metadata"]["statut"] = "brouillon"
        elif review.decision == "rejetee":
            fiche_dict["metadata"]["statut"] = "archivee"

        fiche_dict["metadata"]["date_maj"] = datetime.now()
        fiche_dict["metadata"]["version"] = fiche_dict["metadata"].get("version", 1) + 1

        updated_fiche = FicheMetier(**fiche_dict)
        repo.update_fiche(updated_fiche)

        statut_labels = {
            "approuvee": "publiee",
            "a_corriger": "brouillon",
            "rejetee": "archivee",
        }

        decision_labels = {
            "approuvee": "Approuvée",
            "a_corriger": "Renvoyée en correction",
            "rejetee": "Rejetée",
        }

        user_name = current_user.get("name", "Utilisateur")
        comment_part = f" - Commentaire: {review.commentaire}" if review.commentaire else ""
        log_action(
            TypeEvenement.VALIDATION,
            f"Review de {code_rome} ({fiche.nom_epicene}) par {user_name}: {decision_labels[review.decision]}{comment_part}",
            code_rome=code_rome,
            agent=user_name,
            validateur=user_name,
        )

        return {
            "message": f"Fiche {review.decision}",
            "code_rome": code_rome,
            "decision": review.decision,
            "commentaire": review.commentaire,
            "nouveau_statut": statut_labels[review.decision],
            "publie_par": user_name,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur review: {str(e)}")


# ==================== DELETE ====================

@app.delete("/api/fiches/{code_rome}")
async def delete_fiche(code_rome: str, current_user: dict = Depends(get_current_user)):
    """Supprime définitivement une fiche métier."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        nom = fiche.nom_epicene

        # Supprimer via SQL direct
        from sqlalchemy import delete as sql_delete
        from database.models import FicheMetierDB
        with repo.session() as session:
            session.execute(sql_delete(FicheMetierDB).where(FicheMetierDB.code_rome == code_rome))

        user_name = current_user.get("name", "Utilisateur")
        log_action(
            TypeEvenement.ARCHIVAGE,
            f"Suppression définitive de {code_rome} ({nom}) par {user_name}",
            code_rome=code_rome,
            agent=user_name,
        )

        return {
            "message": f"Fiche {code_rome} supprimée",
            "code_rome": code_rome,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur suppression: {str(e)}")


# ==================== AUTO-CORRECTION IA ====================

AUTOCORRECT_PROMPT = """Tu es un expert en ressources humaines et en fiches métiers ROME.

La fiche métier suivante a été analysée et a obtenu un score de complétude insuffisant.
Tu dois CORRIGER et COMPLÉTER UNIQUEMENT les sections qui sont incomplètes ou manquantes.

Fiche actuelle :
- Code ROME : {code_rome}
- Nom : {nom_masculin} / {nom_feminin}
- Description : {description}
- Description courte : {description_courte}
- Missions principales ({nb_missions} éléments) : {missions_principales}
- Accès métier : {acces_metier}
- Compétences techniques ({nb_comp} éléments) : {competences}
- Compétences transversales ({nb_comp_trans} éléments) : {competences_transversales}
- Savoirs ({nb_savoirs} éléments) : {savoirs}
- Formations ({nb_formations} éléments) : {formations}
- Certifications ({nb_certifs} éléments) : {certifications}
- Conditions de travail ({nb_conditions} éléments) : {conditions_travail}
- Environnements ({nb_envs} éléments) : {environnements}
- Secteurs d'activité ({nb_secteurs} éléments) : {secteurs_activite}
- Salaires : {salaires}
- Perspectives : {perspectives}
- Types de contrats : {types_contrats}
- Mobilité : {mobilite}

PROBLÈMES IDENTIFIÉS :
{problemes}

SUGGESTIONS D'AMÉLIORATION :
{suggestions}

OBJECTIFS MINIMUM PAR SECTION :
- description : 5-8 phrases détaillées
- description_courte : 1 phrase percutante (max 200 car.)
- missions_principales : minimum 8 éléments
- acces_metier : 4-6 phrases détaillées
- competences : minimum 10 éléments
- competences_transversales : minimum 6 éléments
- savoirs : minimum 8 éléments
- formations : minimum 5 éléments
- certifications : minimum 3 éléments
- conditions_travail : minimum 5 éléments
- environnements : minimum 4 éléments
- secteurs_activite : minimum 4 éléments
- salaires, perspectives, types_contrats, mobilite : tous les champs remplis

Réponds UNIQUEMENT avec un objet JSON contenant TOUTES les sections (même celles déjà complètes, pour ne rien perdre).
Le JSON doit avoir exactement la même structure que le prompt d'enrichissement initial.
ENRICHIS les sections faibles en AJOUTANT des éléments pertinents aux listes existantes.
Ne supprime AUCUN élément existant, ajoute seulement.

{{
    "description": "...",
    "description_courte": "...",
    "missions_principales": ["..."],
    "acces_metier": "...",
    "competences": ["..."],
    "competences_transversales": ["..."],
    "savoirs": ["..."],
    "formations": ["..."],
    "certifications": ["..."],
    "conditions_travail": ["..."],
    "environnements": ["..."],
    "secteurs_activite": ["..."],
    "salaires": {{"junior": {{"min": 0, "max": 0, "median": 0}}, "confirme": {{"min": 0, "max": 0, "median": 0}}, "senior": {{"min": 0, "max": 0, "median": 0}}}},
    "perspectives": {{"tension": 0.0, "tendance": "stable", "evolution_5ans": "...", "nombre_offres": 0, "taux_insertion": 0.0}},
    "types_contrats": {{"cdi": 0, "cdd": 0, "interim": 0, "autre": 0}},
    "mobilite": {{"metiers_proches": [{{"nom": "...", "contexte": "..."}}], "evolutions": [{{"nom": "...", "contexte": "..."}}]}}
}}"""


class AutoCorrectRequest(BaseModel):
    problemes: List[str] = []
    suggestions: List[str] = []


@app.post("/api/fiches/{code_rome}/auto-correct")
def auto_correct_fiche(code_rome: str, rapport: AutoCorrectRequest, current_user: dict = Depends(get_current_user)):
    """Correction automatique IA : comble les lacunes identifiées par la validation."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        api_key = os.getenv("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY non configurée")

        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        prompt = AUTOCORRECT_PROMPT.format(
            code_rome=fiche.code_rome,
            nom_masculin=fiche.nom_masculin,
            nom_feminin=fiche.nom_feminin,
            description=fiche.description or "Non renseignée",
            description_courte=fiche.description_courte or "Non renseignée",
            nb_missions=len(fiche.missions_principales or []),
            missions_principales=json.dumps(fiche.missions_principales or [], ensure_ascii=False),
            acces_metier=fiche.acces_metier or "Non renseigné",
            nb_comp=len(fiche.competences or []),
            competences=json.dumps(fiche.competences or [], ensure_ascii=False),
            nb_comp_trans=len(fiche.competences_transversales or []),
            competences_transversales=json.dumps(fiche.competences_transversales or [], ensure_ascii=False),
            nb_savoirs=len(fiche.savoirs or []),
            savoirs=json.dumps(fiche.savoirs or [], ensure_ascii=False),
            nb_formations=len(fiche.formations or []),
            formations=json.dumps(fiche.formations or [], ensure_ascii=False),
            nb_certifs=len(fiche.certifications or []),
            certifications=json.dumps(fiche.certifications or [], ensure_ascii=False),
            nb_conditions=len(fiche.conditions_travail or []),
            conditions_travail=json.dumps(fiche.conditions_travail or [], ensure_ascii=False),
            nb_envs=len(fiche.environnements or []),
            environnements=json.dumps(fiche.environnements or [], ensure_ascii=False),
            nb_secteurs=len(fiche.secteurs_activite or []),
            secteurs_activite=json.dumps(fiche.secteurs_activite or [], ensure_ascii=False),
            salaires=json.dumps(fiche.salaires.model_dump() if fiche.salaires else {}, ensure_ascii=False),
            perspectives=json.dumps(fiche.perspectives.model_dump() if fiche.perspectives else {}, ensure_ascii=False),
            types_contrats=json.dumps(fiche.types_contrats.model_dump() if fiche.types_contrats else {}, ensure_ascii=False),
            mobilite=json.dumps(fiche.mobilite.model_dump() if fiche.mobilite else {}, ensure_ascii=False),
            problemes="\n".join(f"- {p}" for p in rapport.problemes) or "Aucun problème spécifique",
            suggestions="\n".join(f"- {s}" for s in rapport.suggestions) or "Aucune suggestion spécifique",
        )

        response = claude_call_with_retry(
            client,
            model=CLAUDE_MODEL,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )

        content = response.content[0].text.strip()
        json_match = re.search(r'\{[\s\S]*\}', content)
        if not json_match:
            raise HTTPException(status_code=500, detail="Réponse Claude invalide (pas de JSON)")

        data = json.loads(json_match.group())

        # Appliquer les corrections
        fiche_dict = fiche.model_dump()
        for key in [
            "description", "description_courte", "missions_principales",
            "acces_metier", "competences", "competences_transversales",
            "savoirs", "formations", "certifications", "conditions_travail",
            "environnements", "secteurs_activite",
        ]:
            if key in data:
                fiche_dict[key] = data[key]

        if "salaires" in data:
            fiche_dict["salaires"] = data["salaires"]
        if "perspectives" in data:
            fiche_dict["perspectives"] = data["perspectives"]
        if "types_contrats" in data:
            fiche_dict["types_contrats"] = data["types_contrats"]
        if "mobilite" in data:
            fiche_dict["mobilite"] = data["mobilite"]

        fiche_dict["metadata"]["date_maj"] = datetime.now()
        fiche_dict["metadata"]["version"] = fiche_dict["metadata"].get("version", 1) + 1

        updated_fiche = FicheMetier(**fiche_dict)
        repo.update_fiche(updated_fiche)

        user_name = current_user.get("name", "Utilisateur")
        log_action(
            TypeEvenement.CORRECTION,
            f"Correction automatique IA de {code_rome} ({fiche.nom_epicene}) par {user_name} - v{updated_fiche.metadata.version}",
            code_rome=code_rome,
            agent=user_name,
        )

        return {
            "message": "Fiche corrigée automatiquement",
            "code_rome": code_rome,
            "nom": fiche.nom_epicene,
            "version": updated_fiche.metadata.version,
        }

    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"Erreur parsing JSON Claude: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur auto-correction: {str(e)}")


# ==================== GENERATION DE VARIANTES ====================

VARIANTES_PROMPT = """Tu es un expert en adaptation de contenus pédagogiques et multilingues.

FICHE SOURCE :
- Code ROME : {code_rome}
- Nom : {nom_masculin} / {nom_feminin}
- Description : {description}
- Compétences : {competences}
- Formations : {formations}

TÂCHE : Générer EXACTEMENT {nb_variantes} variantes de cette fiche selon les axes suivants :
- Langues : {langues_str}
- Tranches d'âge : {tranches_str}
- Formats : {formats_str}
- Genres : {genres_str}

Chaque variante DOIT être dans la langue demandée. Tout le contenu (nom, description, compétences, formations, certifications, conditions, environnements) doit être traduit et adapté dans cette langue.

RÈGLES PAR AXE :

1. LANGUES
   - "fr" : Français
   - "en" : Anglais (English) — tout le contenu en anglais
   - "es" : Espagnol (Español) — tout le contenu en espagnol
   - "it" : Italien (Italiano) — tout le contenu en italien
   - "de" : Allemand (Deutsch) — tout le contenu en allemand
   - "pt" : Portugais (Português) — tout le contenu en portugais
   - "ar" : Arabe (العربية) — tout le contenu en arabe avec script arabe

2. TRANCHES D'ÂGE
   - "11-15" : Langage simple, exemples concrets, ton encourageant, phrases courtes (<20 mots)
   - "15-18" : Vocabulaire jeune, orientation études, exemples inspirants, phrases moyennes (<25 mots)
   - "18+" : Langage professionnel, exhaustif, technique si nécessaire

3. FORMATS
   - "standard" : Rédaction classique
   - "falc" (Facile À Lire et à Comprendre) : Phrases <15 mots, vocabulaire simple (niveau primaire), 1 idée par phrase, pas de jargon

4. GENRES
   - "masculin" : Utiliser le masculin partout (ou équivalent dans la langue)
   - "feminin" : Utiliser le féminin partout (ou équivalent dans la langue)
   - "epicene" : Langage neutre (éviter les accords genrés)

STRUCTURE DE SORTIE :
Réponds UNIQUEMENT avec un objet JSON valide (sans texte avant ou après) :

{{
    "variantes": [
        {{
            "langue": "{exemple_langue}",
            "tranche_age": "18+",
            "format_contenu": "standard",
            "genre": "masculin",
            "nom": "Nom du métier dans la langue demandée",
            "description": "Description complète dans la langue demandée (3-5 phrases)",
            "description_courte": "Description courte dans la langue demandée (1 phrase max 200 car)",
            "competences": ["Compétence 1 dans la langue", "Compétence 2", "..."],
            "competences_transversales": ["Soft skill 1 dans la langue", "..."],
            "formations": ["Formation 1 dans la langue", "..."],
            "certifications": ["Certification 1", "..."],
            "conditions_travail": ["Condition 1 dans la langue", "..."],
            "environnements": ["Environnement 1 dans la langue", "..."]
        }}
    ]
}}

IMPORTANT :
- Le champ "langue" de chaque variante DOIT correspondre à la langue demandée (ex: "en" pour anglais, "es" pour espagnol)
- Pour FALC : PHRASES <15 MOTS, vocabulaire niveau CM1-CM2
- Pour 11-15 ans : Éviter jargon, expliquer concepts
- Pour genre épicène : Utiliser des tournures neutres
- Pour l'arabe, écrire en script arabe"""


class GenerateVariantesRequest(BaseModel):
    genres: List[str] = ["masculin", "feminin", "epicene"]
    tranches_age: List[str] = ["18+"]
    formats: List[str] = ["standard", "falc"]
    langues: List[str] = ["fr"]


@app.post("/api/fiches/{code_rome}/variantes/generate")
def generate_variantes(code_rome: str, request: GenerateVariantesRequest, current_user: dict = Depends(get_current_user)):
    """Génère des variantes d'une fiche via Claude API (batch par lots de 3 max)."""
    try:
        fiche = repo.get_fiche(code_rome)
        if not fiche:
            raise HTTPException(status_code=404, detail=f"Fiche {code_rome} non trouvée")

        if not fiche.description:
            raise HTTPException(status_code=400, detail="Fiche non enrichie, impossible de générer des variantes")

        api_key = os.getenv("ANTHROPIC_API_KEY", "")
        if not api_key:
            raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY non configurée")

        # Validate inputs
        valid_genres = {"masculin", "feminin", "epicene"}
        valid_tranches = {"18+", "15-18", "11-15"}
        valid_formats = {"standard", "falc"}
        valid_langues = {"fr", "en", "es", "de", "it", "pt", "ar"}

        genres = [g for g in request.genres if g in valid_genres]
        tranches = [t for t in request.tranches_age if t in valid_tranches]
        formats = [f for f in request.formats if f in valid_formats]
        langues = [l for l in request.langues if l in valid_langues] or ["fr"]

        if not genres or not tranches or not formats:
            raise HTTPException(status_code=400, detail="Au moins un genre, une tranche d'âge et un format sont requis")

        # Build all combinations (langue, tranche, format, genre)
        all_combos = []
        for lang in langues:
            for t in tranches:
                for fmt in formats:
                    for g in genres:
                        all_combos.append((lang, t, fmt, g))

        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        # Map string values to enums
        genre_map = {
            "masculin": GenreGrammatical.MASCULIN,
            "feminin": GenreGrammatical.FEMININ,
            "epicene": GenreGrammatical.EPICENE,
        }
        tranche_map = {
            "11-15": TrancheAge.JEUNE_11_15,
            "15-18": TrancheAge.ADOS_15_18,
            "18+": TrancheAge.ADULTE,
        }
        format_map = {
            "standard": FormatContenu.STANDARD,
            "falc": FormatContenu.FALC,
        }
        langue_map = {
            "fr": LangueSupporte.FR,
            "en": LangueSupporte.EN,
            "es": LangueSupporte.ES,
            "de": LangueSupporte.DE,
            "it": LangueSupporte.IT,
            "pt": LangueSupporte.PT,
            "ar": LangueSupporte.AR,
        }

        # Batch: max 3 variantes per Claude call to stay within rate limits (8k output tokens/min)
        BATCH_SIZE = 3
        saved_count = 0

        for i in range(0, len(all_combos), BATCH_SIZE):
            # Wait between batches to respect rate limits
            if i > 0:
                _time.sleep(30)

            batch = all_combos[i:i + BATCH_SIZE]
            batch_langues = list({c[0] for c in batch})
            batch_tranches = list({c[1] for c in batch})
            batch_formats = list({c[2] for c in batch})
            batch_genres = list({c[3] for c in batch})

            prompt = VARIANTES_PROMPT.format(
                code_rome=fiche.code_rome,
                nom_masculin=fiche.nom_masculin,
                nom_feminin=fiche.nom_feminin,
                description=fiche.description or "",
                competences=", ".join((fiche.competences or [])[:5]) + ("..." if len(fiche.competences or []) > 5 else ""),
                formations=", ".join((fiche.formations or [])[:3]) + ("..." if len(fiche.formations or []) > 3 else ""),
                nb_variantes=len(batch),
                tranches_str=", ".join(batch_tranches),
                formats_str=", ".join(batch_formats),
                genres_str=", ".join(batch_genres),
                langues_str=", ".join(batch_langues),
                exemple_langue=batch_langues[0],
            )

            response = claude_call_with_retry(
                client,
                model=CLAUDE_MODEL,
                max_tokens=4096,
                messages=[{"role": "user", "content": prompt}],
            )

            content = response.content[0].text.strip()
            json_match = re.search(r'\{[\s\S]*\}', content)
            if not json_match:
                print(f"Batch {i//BATCH_SIZE + 1}: pas de JSON dans la réponse Claude")
                continue

            try:
                data = json.loads(json_match.group())
            except json.JSONDecodeError as e:
                print(f"Batch {i//BATCH_SIZE + 1}: erreur JSON: {e}")
                continue

            variantes_data = data.get("variantes", [])

            for v_data in variantes_data:
                try:
                    variante = VarianteFiche(
                        code_rome=code_rome,
                        langue=langue_map.get(v_data.get("langue", "fr"), LangueSupporte.FR),
                        tranche_age=tranche_map.get(v_data.get("tranche_age", "18+"), TrancheAge.ADULTE),
                        format_contenu=format_map.get(v_data.get("format_contenu", "standard"), FormatContenu.STANDARD),
                        genre=genre_map.get(v_data.get("genre", "masculin"), GenreGrammatical.MASCULIN),
                        nom=v_data.get("nom", fiche.nom_epicene),
                        description=v_data.get("description", ""),
                        description_courte=v_data.get("description_courte"),
                        competences=v_data.get("competences", []),
                        competences_transversales=v_data.get("competences_transversales", []),
                        formations=v_data.get("formations", []),
                        certifications=v_data.get("certifications", []),
                        conditions_travail=v_data.get("conditions_travail", []),
                        environnements=v_data.get("environnements", []),
                    )
                    repo.save_variante(variante)
                    saved_count += 1
                except Exception as e:
                    print(f"Erreur sauvegarde variante: {e}")

        user_name = current_user.get("name", "Utilisateur")
        log_action(
            TypeEvenement.MODIFICATION,
            f"Génération de {saved_count} variantes pour {code_rome} ({fiche.nom_epicene}) par {user_name}",
            code_rome=code_rome,
            agent=user_name,
        )

        return {
            "message": f"{saved_count} variantes générées pour {code_rome}",
            "code_rome": code_rome,
            "variantes_generees": saved_count,
        }

    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur parsing JSON Claude: {str(e)}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erreur génération variantes: {str(e)}")


# ==================== ROME SYNC ====================

def parse_nom_genre(nom_complet: str) -> dict:
    """Parse un nom au format 'Masculin / Féminin complément' du ROME."""
    if ' / ' not in nom_complet:
        return {'masculin': nom_complet, 'feminin': nom_complet, 'epicene': nom_complet}

    parts = nom_complet.split(' / ', 1)
    masculin_part = parts[0].strip()
    feminin_rest = parts[1].strip()

    masc_words = masculin_part.split()
    fem_words = feminin_rest.split()
    nb_mots_nom = len(masc_words)

    if len(fem_words) > nb_mots_nom:
        fem_nom = ' '.join(fem_words[:nb_mots_nom])
        complement = ' '.join(fem_words[nb_mots_nom:])
        masculin = f"{masculin_part} {complement}"
        feminin = f"{fem_nom} {complement}"
    else:
        masculin = masculin_part
        feminin = feminin_rest

    epicene = nom_complet
    return {'masculin': masculin, 'feminin': feminin, 'epicene': epicene}


@app.post("/api/rome/sync")
async def sync_rome(current_user: dict = Depends(get_current_user)):
    """Synchronise le référentiel ROME depuis data.gouv.fr (télécharge arborescence_principale.xlsx)."""
    import tempfile
    import httpx
    import openpyxl

    try:
        # 1. Récupérer les métadonnées du dataset ROME sur data.gouv.fr
        dataset_url = "https://www.data.gouv.fr/api/1/datasets/repertoire-operationnel-des-metiers-et-des-emplois-rome/"
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            resp = await client.get(dataset_url)
            if resp.status_code != 200:
                raise HTTPException(status_code=502, detail=f"Impossible de contacter data.gouv.fr (HTTP {resp.status_code})")
            dataset = resp.json()

        # 2. Trouver l'URL du fichier arborescence_principale.xlsx
        xlsx_url = None
        for resource in dataset.get("resources", []):
            title = (resource.get("title") or "").lower()
            url = resource.get("url", "")
            if "arborescence" in title and url.endswith(".xlsx"):
                xlsx_url = url
                break

        if not xlsx_url:
            raise HTTPException(status_code=404, detail="Fichier arborescence_principale.xlsx introuvable dans le dataset ROME")

        # 3. Télécharger le fichier XLSX
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            resp = await client.get(xlsx_url)
            if resp.status_code != 200:
                raise HTTPException(status_code=502, detail=f"Impossible de télécharger le XLSX (HTTP {resp.status_code})")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(resp.content)
            tmp_path = tmp.name

        # 4. Parser le fichier XLSX (même logique que import_rome.py)
        try:
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb[wb.sheetnames[1]]  # Deuxième feuille = données

            fiches_rome = {}
            current_domaine_nom = ""
            current_sous_domaine_nom = ""

            for row in ws.iter_rows(min_row=2, values_only=True):
                col1 = str(row[0]).strip() if row[0] else ""
                col2 = str(row[1]).strip() if row[1] else ""
                col3 = str(row[2]).strip() if row[2] else ""
                col4 = str(row[3]).strip() if row[3] else ""

                # Grand domaine
                if col1 and not col2 and not col3 and col4:
                    current_domaine_nom = col4
                    continue

                # Sous-domaine
                if col1 and col2 and not col3 and col4:
                    current_sous_domaine_nom = col4
                    continue

                # Fiche
                if col3 and col1 and col2:
                    code_rome = f"{col1}{col2}{col3}"
                    if code_rome not in fiches_rome:
                        noms = parse_nom_genre(col4)
                        fiches_rome[code_rome] = {
                            'nom_masculin': noms['masculin'],
                            'nom_feminin': noms['feminin'],
                            'nom_epicene': noms['epicene'],
                        }

            wb.close()
        finally:
            os.unlink(tmp_path)

        # 5. Comparer avec la base et créer/mettre à jour
        from database.models import MetadataFiche

        nouvelles = 0
        mises_a_jour = 0
        inchangees = 0

        for code_rome, rome_data in fiches_rome.items():
            existing = repo.get_fiche(code_rome)

            if not existing:
                # Nouvelle fiche → créer en brouillon
                nouvelle_fiche = FicheMetier(
                    id=code_rome,
                    code_rome=code_rome,
                    nom_masculin=rome_data['nom_masculin'],
                    nom_feminin=rome_data['nom_feminin'],
                    nom_epicene=rome_data['nom_epicene'],
                    metadata=MetadataFiche(statut=StatutFiche.BROUILLON, version=1),
                )
                repo.create_fiche(nouvelle_fiche)
                nouvelles += 1
            else:
                # Fiche existante → vérifier si les noms ont changé
                changed = (
                    existing.nom_masculin != rome_data['nom_masculin']
                    or existing.nom_feminin != rome_data['nom_feminin']
                    or existing.nom_epicene != rome_data['nom_epicene']
                )
                if changed:
                    fiche_dict = existing.model_dump()
                    fiche_dict['nom_masculin'] = rome_data['nom_masculin']
                    fiche_dict['nom_feminin'] = rome_data['nom_feminin']
                    fiche_dict['nom_epicene'] = rome_data['nom_epicene']
                    fiche_dict['metadata']['date_maj'] = datetime.now()
                    updated_fiche = FicheMetier(**fiche_dict)
                    repo.update_fiche(updated_fiche)
                    mises_a_jour += 1
                else:
                    inchangees += 1

        user_name = current_user.get("name", "Utilisateur")
        log_action(
            TypeEvenement.MODIFICATION,
            f"Synchronisation ROME par {user_name} : {nouvelles} nouvelles, {mises_a_jour} mises à jour, {inchangees} inchangées",
            agent=user_name,
        )

        return {
            "message": f"Synchronisation ROME terminée",
            "nouvelles": nouvelles,
            "mises_a_jour": mises_a_jour,
            "inchangees": inchangees,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur synchronisation ROME: {str(e)}")


# ==================== REGIONAL DATA (France Travail Offres API) ====================

FRANCE_REGIONS = [
    {"code": "01", "libelle": "Guadeloupe"},
    {"code": "02", "libelle": "Martinique"},
    {"code": "03", "libelle": "Guyane"},
    {"code": "04", "libelle": "La Réunion"},
    {"code": "06", "libelle": "Mayotte"},
    {"code": "11", "libelle": "Île-de-France"},
    {"code": "24", "libelle": "Centre-Val de Loire"},
    {"code": "27", "libelle": "Bourgogne-Franche-Comté"},
    {"code": "28", "libelle": "Normandie"},
    {"code": "32", "libelle": "Hauts-de-France"},
    {"code": "44", "libelle": "Grand Est"},
    {"code": "52", "libelle": "Pays de la Loire"},
    {"code": "53", "libelle": "Bretagne"},
    {"code": "75", "libelle": "Nouvelle-Aquitaine"},
    {"code": "76", "libelle": "Occitanie"},
    {"code": "84", "libelle": "Auvergne-Rhône-Alpes"},
    {"code": "93", "libelle": "Provence-Alpes-Côte d'Azur"},
    {"code": "94", "libelle": "Corse"},
]


def _get_ft_token():
    """Obtient un token OAuth2 France Travail (Offres d'emploi v2)."""
    client_id = os.getenv("FT_CLIENT_ID", "")
    client_secret = os.getenv("FT_CLIENT_SECRET", "")
    if not client_id or not client_secret:
        return None

    import httpx as _httpx
    resp = _httpx.post(
        "https://entreprise.francetravail.fr/connexion/oauth2/access_token?realm=/partenaire",
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "api_offresdemploiv2 o2dsoffre",
        },
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=15.0,
    )
    if resp.status_code != 200:
        return None
    return resp.json().get("access_token")


def _compute_salary_stats(values: list) -> dict | None:
    """Calcule min/max/median/nb_offres depuis une liste de salaires."""
    if not values:
        return None
    values.sort()
    n = len(values)
    return {"min": round(values[0]), "max": round(values[-1]), "median": round(values[n // 2]), "nb_offres": n}


def _parse_salary_annual(libelle: str) -> float | None:
    """Extrait un salaire annuel brut approximatif depuis le libellé France Travail."""
    if not libelle:
        return None
    import re as _re
    nums = [float(n) for n in _re.findall(r'(\d+(?:\.\d+)?)', libelle) if float(n) > 100]
    if not nums:
        return None

    if "Annuel" in libelle or "Mensuel" in libelle:
        # Values > 1000 are likely annual; take average of first two
        vals = [n for n in nums if n > 1000]
        if vals:
            return sum(vals[:2]) / len(vals[:2])
    elif "Horaire" in libelle:
        vals = [n for n in nums if n < 100]
        if vals:
            avg = sum(vals[:2]) / len(vals[:2])
            return avg * 1607  # heures/an en France
    return None


@app.get("/api/regions")
async def get_regions():
    """Retourne la liste des régions françaises."""
    return {"regions": FRANCE_REGIONS}


@app.get("/api/fiches/{code_rome}/regional")
async def get_regional_data(
    code_rome: str,
    region: str = Query(..., description="Code région (ex: 11 pour Île-de-France)"),
):
    """Récupère les données régionales pour un métier via l'API Offres d'emploi France Travail."""
    import httpx as _httpx

    token = _get_ft_token()
    if not token:
        raise HTTPException(
            status_code=503,
            detail="Credentials France Travail non configurés (FT_CLIENT_ID / FT_CLIENT_SECRET)"
        )

    # Vérifier que la région existe
    region_name = next((r["libelle"] for r in FRANCE_REGIONS if r["code"] == region), None)
    if not region_name:
        raise HTTPException(status_code=400, detail=f"Région inconnue: {region}")

    try:
        async with _httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.get(
                "https://api.francetravail.io/partenaire/offresdemploi/v2/offres/search",
                params={"codeROME": code_rome, "region": region, "range": "0-149"},
                headers={"Authorization": f"Bearer {token}"},
            )

        if resp.status_code in (204, 404):
            return {
                "region": region,
                "region_name": region_name,
                "code_rome": code_rome,
                "nb_offres": 0,
                "salaires": None,
                "types_contrats": None,
                "salaires_par_niveau": None,
                "experience_distribution": None,
                "tension_regionale": None,
            }

        if resp.status_code not in (200, 206):
            raise HTTPException(status_code=502, detail=f"Erreur API France Travail: HTTP {resp.status_code}")

        data = resp.json()
        resultats = data.get("resultats", [])
        filtres = data.get("filtresPossibles", [])

        # Nombre total d'offres depuis les filtres (plus précis que len(resultats))
        nb_offres = 0
        for f in filtres:
            if f["filtre"] == "typeContrat":
                nb_offres = sum(a["nbResultats"] for a in f["agregation"])
                break
        if nb_offres == 0:
            nb_offres = len(resultats)

        # Agréger les salaires (global + par niveau d'expérience)
        import re as _re_sal
        salaires_tous = []
        salaires_par_exp = {"junior": [], "confirme": [], "senior": []}

        for r in resultats:
            sal = _parse_salary_annual(r.get("salaire", {}).get("libelle", ""))
            if sal and 15000 < sal < 200000:
                salaires_tous.append(sal)
                exp = r.get("experienceExige", "")
                exp_lib = r.get("experienceLibelle", "")
                if exp == "D":
                    salaires_par_exp["junior"].append(sal)
                elif exp == "E":
                    years_match = _re_sal.search(r'(\d+)\s*An', exp_lib)
                    years = int(years_match.group(1)) if years_match else 3
                    if years <= 3:
                        salaires_par_exp["confirme"].append(sal)
                    else:
                        salaires_par_exp["senior"].append(sal)
                elif exp == "S":
                    salaires_par_exp["confirme"].append(sal)

        salaires_tous.sort()
        salaires_result = None
        if salaires_tous:
            n = len(salaires_tous)
            salaires_result = {
                "nb_offres_avec_salaire": n,
                "min": round(salaires_tous[0]),
                "max": round(salaires_tous[-1]),
                "median": round(salaires_tous[n // 2]),
                "moyenne": round(sum(salaires_tous) / n),
            }

        # Salaires par niveau d'expérience
        salaires_par_niveau = None
        j = _compute_salary_stats(salaires_par_exp["junior"])
        c = _compute_salary_stats(salaires_par_exp["confirme"])
        s = _compute_salary_stats(salaires_par_exp["senior"])
        if j or c or s:
            salaires_par_niveau = {"junior": j, "confirme": c, "senior": s}

        # Agréger les types de contrats
        contrats_result = None
        for f in filtres:
            if f["filtre"] == "typeContrat":
                cdi = cdd = interim = autre = 0
                for a in f["agregation"]:
                    v = a["valeurPossible"]
                    nb = a["nbResultats"]
                    if v == "CDI":
                        cdi = nb
                    elif v == "CDD":
                        cdd = nb
                    elif v == "MIS":
                        interim = nb
                    else:
                        autre += nb
                total = cdi + cdd + interim + autre
                if total > 0:
                    contrats_result = {
                        "cdi": round(cdi / total * 100),
                        "cdd": round(cdd / total * 100),
                        "interim": round(interim / total * 100),
                        "autre": round(autre / total * 100),
                    }
                break

        # Distribution d'expérience depuis filtresPossibles
        experience_dist = None
        for f in filtres:
            if f["filtre"] == "experience":
                junior_count = confirme_count = senior_count = 0
                for a in f["agregation"]:
                    v, nb = a["valeurPossible"], a["nbResultats"]
                    if v in ("0", "4"):
                        junior_count += nb
                    elif v == "2":
                        confirme_count += nb
                    elif v == "3":
                        senior_count += nb
                total_exp = junior_count + confirme_count + senior_count
                if total_exp > 0:
                    experience_dist = {
                        "junior": junior_count, "confirme": confirme_count, "senior": senior_count,
                        "junior_pct": round(junior_count / total_exp * 100),
                        "confirme_pct": round(confirme_count / total_exp * 100),
                        "senior_pct": round(senior_count / total_exp * 100),
                    }
                break

        # Tension régionale (proxy basé sur CDI ratio, salaire moyen, densité d'offres)
        tension_regionale = None
        if nb_offres > 0:
            cdi_ratio = contrats_result["cdi"] / 100.0 if contrats_result else 0.5
            salary_score = 0.5
            if salaires_tous:
                salary_score = max(0.0, min(1.0, (sum(salaires_tous) / len(salaires_tous) - 25000) / 30000))
            density_score = min(1.0, nb_offres / 150.0)
            tension_regionale = round(cdi_ratio * 0.4 + salary_score * 0.3 + density_score * 0.3, 2)

        return {
            "region": region,
            "region_name": region_name,
            "code_rome": code_rome,
            "nb_offres": nb_offres,
            "salaires": salaires_result,
            "types_contrats": contrats_result,
            "salaires_par_niveau": salaires_par_niveau,
            "experience_distribution": experience_dist,
            "tension_regionale": tension_regionale,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur données régionales: {str(e)}")


@app.get("/api/fiches/{code_rome}/recrutements")
async def get_recrutements_par_mois(
    code_rome: str,
    region: str = Query(None, description="Code région (optionnel, ex: 11 pour Île-de-France)"),
):
    """Récupère le nombre d'offres publiées par mois (12 derniers mois) via l'API Offres France Travail."""
    import httpx as _httpx
    import asyncio
    from datetime import date
    from dateutil.relativedelta import relativedelta

    token = _get_ft_token()
    if not token:
        raise HTTPException(
            status_code=503,
            detail="Credentials France Travail non configurés (FT_CLIENT_ID / FT_CLIENT_SECRET)"
        )

    region_name = None
    if region:
        region_name = next((r["libelle"] for r in FRANCE_REGIONS if r["code"] == region), None)
        if not region_name:
            raise HTTPException(status_code=400, detail=f"Région inconnue: {region}")

    # Build list of 12 months (from 11 months ago to current month)
    today = date.today()
    months = []
    for i in range(11, -1, -1):
        d = today - relativedelta(months=i)
        first_day = d.replace(day=1)
        # Last day of month
        next_month = first_day + relativedelta(months=1)
        last_day = next_month - relativedelta(days=1)
        months.append({
            "label": f"{first_day.strftime('%Y-%m')}",
            "start": f"{first_day.isoformat()}T00:00:00Z",
            "end": f"{last_day.isoformat()}T23:59:59Z",
        })

    async def fetch_month_count(client: _httpx.AsyncClient, month: dict) -> dict:
        params = {
            "codeROME": code_rome,
            "minCreationDate": month["start"],
            "maxCreationDate": month["end"],
            "range": "0-0",
        }
        if region:
            params["region"] = region
        try:
            resp = await client.get(
                "https://api.francetravail.io/partenaire/offresdemploi/v2/offres/search",
                params=params,
                headers={"Authorization": f"Bearer {token}"},
            )
            if resp.status_code in (200, 206):
                content_range = resp.headers.get("Content-Range", "")
                total = 0
                if "/" in content_range:
                    try:
                        total = int(content_range.split("/")[1])
                    except (ValueError, IndexError):
                        total = 0
                return {"mois": month["label"], "nb_offres": total}
            elif resp.status_code == 204:
                return {"mois": month["label"], "nb_offres": 0}
            else:
                return {"mois": month["label"], "nb_offres": 0}
        except Exception:
            return {"mois": month["label"], "nb_offres": 0}

    try:
        async with _httpx.AsyncClient(timeout=30.0) as client:
            recrutements = []
            # Batch of 3 calls with 400ms delay to respect rate limit (3 req/s)
            for i in range(0, len(months), 3):
                batch = months[i:i+3]
                results = await asyncio.gather(*[fetch_month_count(client, m) for m in batch])
                recrutements.extend(results)
                if i + 3 < len(months):
                    await asyncio.sleep(0.4)

        return {
            "code_rome": code_rome,
            "region": region,
            "region_name": region_name,
            "recrutements": recrutements,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur recrutements: {str(e)}")


# ==================== OFFRES D'EMPLOI CACHE ====================

OFFRES_CACHE_TTL_HOURS = 24  # Durée de vie du cache en heures


def _get_cached_offres(code_rome: str, region: str | None) -> list | None:
    """Retourne les offres en cache si elles sont encore fraîches, sinon None."""
    from sqlalchemy import text
    try:
        with repo.engine.connect() as conn:
            region_clause = "AND region = :region" if region else "AND region IS NULL"
            row = conn.execute(text(f"""
                SELECT fetched_at FROM offres_cache
                WHERE code_rome = :code_rome {region_clause}
                LIMIT 1
            """), {"code_rome": code_rome, "region": region}).fetchone()

            if not row:
                return None

            fetched_at = row[0]
            age_hours = (datetime.now() - fetched_at).total_seconds() / 3600
            if age_hours > OFFRES_CACHE_TTL_HOURS:
                return None

            rows = conn.execute(text(f"""
                SELECT offre_id, titre, entreprise, lieu, type_contrat, salaire,
                       experience, date_publication, url
                FROM offres_cache
                WHERE code_rome = :code_rome {region_clause}
                ORDER BY date_publication DESC
            """), {"code_rome": code_rome, "region": region}).fetchall()

            return [
                {
                    "offre_id": r[0], "titre": r[1], "entreprise": r[2],
                    "lieu": r[3], "type_contrat": r[4], "salaire": r[5],
                    "experience": r[6],
                    "date_publication": r[7].isoformat() if r[7] else None,
                    "url": r[8],
                }
                for r in rows
            ]
    except Exception:
        return None


def _save_offres_cache(code_rome: str, region: str | None, offres: list):
    """Remplace les offres en cache pour un code_rome + region."""
    from sqlalchemy import text
    try:
        with repo.engine.begin() as conn:
            region_clause = "AND region = :region" if region else "AND region IS NULL"
            conn.execute(text(f"""
                DELETE FROM offres_cache
                WHERE code_rome = :code_rome {region_clause}
            """), {"code_rome": code_rome, "region": region})

            for o in offres:
                conn.execute(text("""
                    INSERT INTO offres_cache
                        (code_rome, region, offre_id, titre, entreprise, lieu,
                         type_contrat, salaire, experience, date_publication, url, fetched_at)
                    VALUES
                        (:code_rome, :region, :offre_id, :titre, :entreprise, :lieu,
                         :type_contrat, :salaire, :experience, :date_publication, :url, NOW())
                """), {
                    "code_rome": code_rome,
                    "region": region,
                    "offre_id": o.get("offre_id", ""),
                    "titre": o.get("titre", ""),
                    "entreprise": o.get("entreprise"),
                    "lieu": o.get("lieu"),
                    "type_contrat": o.get("type_contrat"),
                    "salaire": o.get("salaire"),
                    "experience": o.get("experience"),
                    "date_publication": o.get("date_publication"),
                    "url": o.get("url"),
                })

            # Cleanup : supprimer les entrées de plus de 30 jours (toutes fiches)
            conn.execute(text("""
                DELETE FROM offres_cache
                WHERE fetched_at < NOW() - INTERVAL '30 days'
            """))
    except Exception as e:
        print(f"Save offres cache error: {e}")


async def _fetch_offres_from_api(code_rome: str, region: str | None, token: str) -> list:
    """Récupère les offres depuis l'API France Travail Offres v2."""
    import httpx as _httpx

    params = {"codeROME": code_rome, "range": "0-49"}
    if region:
        params["region"] = region

    async with _httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.get(
            "https://api.francetravail.io/partenaire/offresdemploi/v2/offres/search",
            params=params,
            headers={"Authorization": f"Bearer {token}"},
        )

    if resp.status_code in (204, 404):
        return []
    if resp.status_code not in (200, 206):
        return []

    data = resp.json()
    resultats = data.get("resultats", [])

    offres = []
    for r in resultats:
        # Extraire le salaire lisible
        sal = r.get("salaire", {})
        salaire_txt = sal.get("libelle") or sal.get("commentaire")

        # URL de l'offre
        url = None
        origine = r.get("origineOffre", {})
        url = origine.get("urlOrigine")
        if not url:
            offre_id = r.get("id", "")
            if offre_id:
                url = f"https://candidat.francetravail.fr/offres/recherche/detail/{offre_id}"

        # Date de publication
        date_pub = None
        date_str = r.get("dateCreation")
        if date_str:
            try:
                date_pub = datetime.fromisoformat(date_str.replace("Z", "+00:00")).replace(tzinfo=None)
            except (ValueError, AttributeError):
                pass

        offres.append({
            "offre_id": r.get("id", ""),
            "titre": r.get("intitule", ""),
            "entreprise": r.get("entreprise", {}).get("nom"),
            "lieu": r.get("lieuTravail", {}).get("libelle"),
            "type_contrat": r.get("typeContratLibelle"),
            "salaire": salaire_txt,
            "experience": r.get("experienceLibelle"),
            "date_publication": date_pub,
            "url": url,
        })

    return offres


@app.get("/api/fiches/{code_rome}/offres")
async def get_offres_emploi(
    code_rome: str,
    region: str = Query(None, description="Code région (optionnel)"),
    limit: int = Query(20, ge=1, le=50),
):
    """Récupère les offres d'emploi actuelles pour un métier, avec cache intelligent."""
    # 1. Vérifier le cache
    cached = _get_cached_offres(code_rome, region)
    if cached is not None:
        region_name = None
        if region:
            region_name = next((r["libelle"] for r in FRANCE_REGIONS if r["code"] == region), None)
        return {
            "code_rome": code_rome,
            "region": region,
            "region_name": region_name,
            "total": len(cached),
            "offres": cached[:limit],
            "from_cache": True,
        }

    # 2. Pas de cache frais → fetch depuis l'API
    token = _get_ft_token()
    if not token:
        raise HTTPException(
            status_code=503,
            detail="Credentials France Travail non configurés"
        )

    region_name = None
    if region:
        region_name = next((r["libelle"] for r in FRANCE_REGIONS if r["code"] == region), None)
        if not region_name:
            raise HTTPException(status_code=400, detail=f"Région inconnue: {region}")

    try:
        offres = await _fetch_offres_from_api(code_rome, region, token)

        # 3. Sauvegarder en cache
        _save_offres_cache(code_rome, region, offres)

        return {
            "code_rome": code_rome,
            "region": region,
            "region_name": region_name,
            "total": len(offres),
            "offres": offres[:limit],
            "from_cache": False,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur offres: {str(e)}")


# ==================== AUTH ENDPOINTS ====================

@app.post("/api/auth/register")
async def register(user_data: UserCreate):
    """Creer un nouveau compte utilisateur."""
    from sqlalchemy.orm import Session
    session = Session(repo.engine)
    try:
        # Verifier si l'email existe deja
        existing = session.execute(
            sa_text("SELECT id FROM users WHERE email = :email"),
            {"email": user_data.email}
        ).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail="Cet email est deja utilise")

        # Creer l'utilisateur
        hashed = hash_password(user_data.password)
        session.execute(
            sa_text("INSERT INTO users (email, hashed_password, name) VALUES (:email, :hashed_password, :name)"),
            {"email": user_data.email, "hashed_password": hashed, "name": user_data.name}
        )
        session.commit()

        return {"message": "Compte cree avec succes"}

    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur inscription: {str(e)}")
    finally:
        session.close()


@app.post("/api/auth/login")
async def login(user_data: UserLogin):
    """Connecter un utilisateur existant."""
    from sqlalchemy.orm import Session
    session = Session(repo.engine)
    try:
        user = session.execute(
            sa_text("SELECT id, email, name, hashed_password FROM users WHERE email = :email"),
            {"email": user_data.email}
        ).fetchone()

        if not user or not verify_password(user_data.password, user[3] if user else ""):
            raise HTTPException(status_code=401, detail="Identifiants incorrects")

        token = create_token(user[0], user[1], user[2])
        return {"token": token, "user": {"id": user[0], "email": user[1], "name": user[2]}}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur connexion: {str(e)}")
    finally:
        session.close()


@app.get("/api/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    """Retourne les informations de l'utilisateur connecte."""
    return current_user


class ResetPasswordRequest(BaseModel):
    email: str
    new_password: str
    admin_key: str


@app.post("/api/admin/reset-password")
async def admin_reset_password(data: ResetPasswordRequest):
    """Reset un mot de passe utilisateur (protege par ADMIN_KEY)."""
    if data.admin_key != ADMIN_KEY:
        raise HTTPException(status_code=403, detail="Cle admin invalide")

    from sqlalchemy.orm import Session
    session = Session(repo.engine)
    try:
        user = session.execute(
            sa_text("SELECT id, email FROM users WHERE email = :email"),
            {"email": data.email}
        ).fetchone()

        if not user:
            raise HTTPException(status_code=404, detail=f"Utilisateur {data.email} non trouve")

        new_hash = hash_password(data.new_password)
        session.execute(
            sa_text("UPDATE users SET hashed_password = :pwd WHERE email = :email"),
            {"pwd": new_hash, "email": data.email}
        )
        session.commit()

        return {"message": "Mot de passe mis a jour", "email": data.email}

    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur reset: {str(e)}")
    finally:
        session.close()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)
